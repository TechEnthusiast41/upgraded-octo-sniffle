"""
全省基础数据表格转换与增幅计算工具

该程序用于从全省基础数据表格中提取忻州市的数据，计算同比增幅，并生成三张Excel表格：
1. 今年数据表
2. 去年数据表
3. 包含同比增幅的对比表

作者: 赵圳楠
日期: 2025-05-28
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
import os
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import webbrowser

# --- 全局配置与常量定义 ---

# 默认在读取Excel文件时，在文件顶部跳过的行数
ROWS_TO_SKIP_DEFAULT = 0  # 假设用户提供的"跳过行数"是指数据表格之前完全无关的行

# 定义输出Excel文件中"市"和"县"这两列的固定列名
OUTPUT_CITY_COL_NAME = "地市"
OUTPUT_COUNTY_COL_NAME = "县级市、区"

# 定义GUI主题颜色
PRIMARY_COLOR = "#1E3F66"  # 深蓝色，政务系统常用色
SECONDARY_COLOR = "#2E5984"  # 次要蓝色
ACCENT_COLOR = "#F5F5F5"  # 浅灰色背景
TEXT_COLOR = "#333333"  # 深灰色文本
BUTTON_COLOR = "#2E5984"  # 按钮颜色
BUTTON_TEXT_COLOR = "white"  # 按钮文字颜色
HIGHLIGHT_COLOR = "#4A7AAC"  # 高亮颜色
ERROR_COLOR = "#D32F2F"  # 错误颜色
SUCCESS_COLOR = "#388E3C"  # 成功颜色

# --- 辅助函数 ---

# 全局变量，用于存储用户选择的负指标
negative_indicators = []

def select_negative_indicators(indicators, parent_window):
    """
    弹出一个对话框让用户选择哪些指标是负指标。
    返回用户选择的负指标列表。
    """
    global negative_indicators
    negative_indicators = [] # 重置

    dialog = tk.Toplevel(parent_window)
    dialog.title("选择负指标")
    dialog.geometry("400x500")
    dialog.transient(parent_window) # 设置为模态
    dialog.grab_set() # 禁用主窗口

    # 居中对话框
    dialog.update_idletasks()
    x = parent_window.winfo_x() + (parent_window.winfo_width() // 2) - (dialog.winfo_width() // 2)
    y = parent_window.winfo_y() + (parent_window.winfo_height() // 2) - (dialog.winfo_height() // 2)
    dialog.geometry(f"+{x}+{y}")

    label = tk.Label(dialog, text="请勾选负向指标（越小越好，需要乘以-1）：", font=("Microsoft YaHei UI", 10, "bold"), wraplength=380, justify="left", bg=ACCENT_COLOR, fg=TEXT_COLOR)
    label.pack(pady=10)

    canvas = tk.Canvas(dialog, borderwidth=0, background=ACCENT_COLOR)
    frame = tk.Frame(canvas, background=ACCENT_COLOR)
    vsb = tk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)

    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    canvas.create_window((4,4), window=frame, anchor="nw",
                                  tags="frame")

    # 允许滚动的配置
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    frame.bind("<Configure>", on_frame_configure)

    indicator_vars = {}
    for indicator in indicators:
        var = tk.BooleanVar(value=False) # 默认都是正指标
        cb = tk.Checkbutton(frame, text=indicator, variable=var, font=("Microsoft YaHei UI", 9), bg=ACCENT_COLOR, fg=TEXT_COLOR, anchor="w", selectcolor="white")
        cb.pack(fill="x", padx=5, pady=2)
        indicator_vars[indicator] = var

    def on_confirm():
        global negative_indicators
        for indicator, var in indicator_vars.items():
            if var.get():
                negative_indicators.append(indicator)
        dialog.destroy()

    confirm_button = tk.Button(dialog, text="确认选择", command=on_confirm,
                               bg=PRIMARY_COLOR, fg=BUTTON_TEXT_COLOR,
                               activebackground=SECONDARY_COLOR,
                               font=("Microsoft YaHei UI", 10, "bold"))
    confirm_button.pack(pady=10)

    parent_window.wait_window(dialog) # 等待对话框关闭

    return negative_indicators


# --- 核心数据处理函数 ---

def extract_indicators_from_template(file_path, header_row_idx=1, skip_rows=0):
    """
    从模板Excel文件中提取所有指标名称和对应的列索引
    
    参数:
        file_path (str): Excel文件路径
        header_row_idx (int): 指标名称所在行索引，默认为1
        skip_rows (int): 跳过的行数，默认为0
        
    返回:
        list: 指标名称列表
        list: 数值列索引列表
    """
    try:
        # 读取Excel文件，不指定header
        df = pd.read_excel(file_path, header=None, skiprows=skip_rows)
        
        # 提取指标名称和对应的列索引
        indicators = []
        value_columns = []
        
        # 从指标行提取指标名称
        # 确保 header_row_idx 不超出 DataFrame 的行数范围
        if header_row_idx >= df.shape[0]:
            return [], [] # 如果指标行索引超出范围，则没有指标

        indicator_row = df.iloc[header_row_idx]
        
        # 从第5列开始（索引4），每隔2列提取一个指标
        for i in range(4, df.shape[1], 2):
            indicator = indicator_row[i]
            if pd.notna(indicator):
                # 替换换行符，确保指标名称格式正确
                indicators.append(str(indicator).replace('\n', ''))
                value_columns.append(i)
        
        return indicators, value_columns
    except Exception as e:
        print(f"提取指标时出错: {e}")
        return [], []

def process_excel_file(input_path, output_path, city_filter='忻州', file_type_label='', 
                       rows_to_skip=0, status_var=None, root=None, log_func=None):
    """
    处理Excel文件，提取指定城市的数据并生成新的Excel表格
    
    参数:
        input_path (str): 输入Excel文件路径
        output_path (str): 输出Excel文件路径
        city_filter (str): 要提取的城市名称，默认为'忻州'
        file_type_label (str): 文件类型标签（去年/今年）
        rows_to_skip (int): 要跳过的行数
        status_var (tk.StringVar): 状态栏变量
        root (tk.Tk): Tkinter根窗口
        log_func (function): 日志记录函数
        
    返回:
        pd.DataFrame: 处理后的DataFrame
        list: 提取的指标名称列表
    """
    try:
        # 更新状态
        message = f"正在处理{file_type_label}文件: {os.path.basename(input_path)}..."
        if status_var:
            status_var.set(message)
            if root: root.update_idletasks()
        if log_func: log_func(message)
        
        # 读取Excel文件
        df_input = pd.read_excel(input_path, header=None, skiprows=rows_to_skip)
        
        # 从输入文件中提取指标名称和列索引
        indicators, value_columns = extract_indicators_from_template(input_path, 1, rows_to_skip)
        
        if not indicators:
            error_msg = f"错误: 在{file_type_label}文件中未找到有效指标"
            if status_var:
                status_var.set(error_msg)
            if log_func: log_func(error_msg, level="ERROR")
            messagebox.showerror("指标错误", f"在文件 {input_path} 中未找到有效指标")
            return None, []
        
        # 更新状态
        message = f"正在提取{file_type_label}文件中的{city_filter}市数据..."
        if status_var:
            status_var.set(message)
            if root: root.update_idletasks()
        if log_func: log_func(message)
        
        # 提取城市数据
        city_data = []
        current_city = None
        
        for row_idx in range(df_input.shape[0]):
            # 获取城市名称（C列，索引2）
            city_value = df_input.iloc[row_idx, 2] if df_input.shape[1] > 2 else None
            city_name = str(city_value).strip() if pd.notna(city_value) and str(city_value).strip() != "" else ""
            
            if city_name:
                current_city = city_name
            
            # 检查是否为目标城市
            if current_city == city_filter:
                # 获取县区名称（D列，索引3）
                county_value = df_input.iloc[row_idx, 3] if df_input.shape[1] > 3 else None
                county_name = str(county_value).strip() if pd.notna(county_value) else ""
                
                if not county_name:
                    continue
                
                # 提取该县区的所有指标值
                row_data = [city_filter, county_name]
                
                # 按照模板表格的列顺序提取数据
                for value_col in value_columns:
                    # 提取数值
                    if value_col < df_input.shape[1]:
                        value = df_input.iloc[row_idx, value_col]
                        # 尝试转换为数值，非数值设为NaN
                        value = pd.to_numeric(value, errors='coerce')
                        row_data.append(value)  # 保持NaN，后续统一处理填充
                    else:
                        row_data.append(np.nan)  # 使用NaN代替缺失值
                
                city_data.append(row_data)
        
        if not city_data:
            error_msg = f"错误: 在{file_type_label}文件中未找到{city_filter}市的数据"
            if status_var:
                status_var.set(error_msg)
            if log_func: log_func(error_msg, level="ERROR")
            messagebox.showerror("数据错误", f"在文件 {input_path} 中未找到{city_filter}市的数据")
            return None, []
        
        # 更新状态
        message = f"正在生成{file_type_label}输出表格..."
        if status_var:
            status_var.set(message)
            if root: root.update_idletasks()
        if log_func: log_func(message)
        
        # 创建输出表格的列名
        columns = [OUTPUT_CITY_COL_NAME, OUTPUT_COUNTY_COL_NAME] + indicators
        
        # 创建输出DataFrame
        df_output = pd.DataFrame(city_data, columns=columns)
        
        # 保存到Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 写入自定义表头
            df_header1 = pd.DataFrame([["基础数据"]])
            df_header1.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False, header=False)
            
            df_header2 = pd.DataFrame([["单位："]])
            df_header2.to_excel(writer, sheet_name='Sheet1', startrow=1, index=False, header=False)
            
            # 写入数据
            df_output.to_excel(writer, sheet_name='Sheet1', startrow=2, index=False)
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # 合并表头单元格
            num_cols = len(df_output.columns)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            
            # 调整列宽
            for col_idx, column in enumerate(df_output.columns, 1):
                column_letter = get_column_letter(col_idx)
                # 获取列标题长度
                header_length = len(str(column))
                
                # 获取列数据的最大长度
                max_data_length = 0
                for value in df_output[column]:
                    value_str = str(value) if pd.notna(value) else ""
                    max_data_length = max(max_data_length, len(value_str))
                
                # 设置列宽
                width = max(header_length, max_data_length) + 2
                worksheet.column_dimensions[column_letter].width = width
        
        success_msg = f"成功: {file_type_label}数据已保存到 {os.path.basename(output_path)}"
        if status_var:
            status_var.set(success_msg)
        if log_func: log_func(success_msg, level="SUCCESS")
        
        return df_output, indicators
        
    except Exception as e:
        error_msg = f"错误: 处理{file_type_label}文件失败: {e}"
        if status_var:
            status_var.set(error_msg)
        if log_func: log_func(error_msg, level="ERROR")
        messagebox.showerror("处理错误", f"处理文件 {input_path} 时发生错误:\n{e}")
        return None, []

def extract_indicators_and_counties(df_last_year, df_this_year):
    """
    提取两个DataFrame中的指标名称和县级市名称
    
    参数:
    df_last_year: 去年的数据DataFrame
    df_this_year: 今年的数据DataFrame
    
    返回:
    indicators: 指标名称列表
    counties: 县级市名称列表
    """
    # 提取指标名称（从列名中获取，跳过前两列"地市"和"县级市、区"）
    # 确保列名是字符串类型，以便正确过滤
    indicators = [col for col in df_last_year.columns if col not in [OUTPUT_CITY_COL_NAME, OUTPUT_COUNTY_COL_NAME]]
    
    # 提取县级市名称（从"县级市、区"列获取）
    counties = df_last_year[OUTPUT_COUNTY_COL_NAME].tolist()
    
    return indicators, counties

def calculate_growth_rates(df_last_year, df_this_year, indicators, counties):
    """
    计算每个指标的同期和同比增幅
    
    参数:
    df_last_year: 去年的数据DataFrame
    df_this_year: 今年的数据DataFrame
    indicators: 指标名称列表
    counties: 县级市名称列表
    
    返回:
    result_data: 包含所有指标及其同期和同比增幅的字典
    """
    # 创建结果字典
    result_data = {}
    
    # 遍历每个指标
    for indicator in indicators:
        # 创建该指标的数据字典
        indicator_data = {
            'name': indicator,  # 指标名称
            'current_values': {},  # 今年的值
            'previous_values': {},  # 去年的值（同期）
            'growth_rates': {}  # 同比增幅
        }
        
        # 遍历每个县级市
        for county in counties:
            # 获取去年和今年的值
            # 通过县级市名称和指标名称定位具体的值
            last_year_row = df_last_year[df_last_year[OUTPUT_COUNTY_COL_NAME] == county]
            this_year_row = df_this_year[df_this_year[OUTPUT_COUNTY_COL_NAME] == county]
            
            if not last_year_row.empty and not this_year_row.empty:
                # 确保取出的值是数值类型，如果不是则转为 NaN
                last_year_value = pd.to_numeric(last_year_row[indicator].values[0], errors='coerce')
                this_year_value = pd.to_numeric(this_year_row[indicator].values[0], errors='coerce')
                
                # 存储当前值和同期值
                indicator_data['current_values'][county] = this_year_value if pd.notna(this_year_value) else 0
                indicator_data['previous_values'][county] = last_year_value if pd.notna(last_year_value) else 0
                
                # 计算同比增幅
                if pd.notna(last_year_value) and last_year_value != 0:
                    growth_rate = (this_year_value - last_year_value) / last_year_value * 100
                elif pd.notna(this_year_value) and this_year_value != 0 and (pd.isna(last_year_value) or last_year_value == 0):
                    growth_rate = np.inf # 去年为0或N/A，今年有值，视为无限增长
                else:
                    growth_rate = np.nan # 去年和今年都为0或N/A，或者其他无法计算的情况
                
                indicator_data['growth_rates'][county] = growth_rate
        
        # 将该指标的数据添加到结果字典
        result_data[indicator] = indicator_data
    
    return result_data

def generate_output_table(result_data, counties, output_file):
    """
    生成输出表格，格式与图片样例一致
    
    参数:
    result_data: 包含所有指标及其同期和同比增幅的字典
    counties: 县级市名称列表
    output_file: 输出文件路径
    
    返回:
    df_output: 生成的DataFrame
    """
    # 创建一个空的DataFrame作为输出表格
    # 列为县级市名称，加上一个指标名称列
    columns = ['指标名称'] + counties
    
    # 创建空的行列表
    rows = []
    
    # 遍历每个指标，为每个指标创建三行（当前值、同期值、同比增幅）
    for indicator_name, indicator_data in result_data.items():
        # 当前值行
        current_row = [indicator_name.split('（')[0] if '（' in indicator_name else indicator_name]  # 去掉单位部分作为指标名称
        for county in counties:
            value = indicator_data['current_values'].get(county)
            current_row.append(f"{value:,.2f}" if pd.notna(value) else "N/A")
        rows.append(current_row)
        
        # 同期行
        previous_row = ['同期']
        for county in counties:
            value = indicator_data['previous_values'].get(county)
            previous_row.append(f"{value:,.2f}" if pd.notna(value) else "N/A")
        rows.append(previous_row)
        
        # 同比增幅行
        growth_row = ['同比增幅%']
        for county in counties:
            growth_rate = indicator_data['growth_rates'].get(county)
            if pd.notna(growth_rate):
                if growth_rate == np.inf:
                    growth_row.append("无限增长")
                else:
                    growth_row.append(f"{growth_rate:,.2f}%")
            else:
                growth_row.append("N/A")
        rows.append(growth_row)
    
    # 创建DataFrame
    df_output = pd.DataFrame(rows, columns=columns)
    
    # 保存到Excel文件
    df_output.to_excel(output_file, index=False)
    
    return df_output

def format_excel_with_style(output_file):
    """
    为输出的Excel文件添加样式，使其更接近图片样例
    
    参数:
    output_file: 输出文件路径
    """
    try:
        # 加载工作簿
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 定义样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        for col in range(2, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[col_letter].width = 10
        
        # 应用样式到所有单元格
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # 为指标行和同比增幅行添加黄色背景
                if row % 3 == 1 or row % 3 == 0:  # 指标行和同比增幅行
                    cell.fill = yellow_fill
                
                # 为表头添加粗体
                if row == 1:
                    cell.font = bold_font
        
        # 保存样式化的工作簿
        wb.save(output_file)
        print(f"已为 {output_file} 添加样式")
        
    except Exception as e:
        print(f"添加样式时出错: {e}")

# --- 日志记录函数 ---

def create_log_function(log_text):
    """创建日志记录函数"""
    def log_message(message, level="INFO"):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        level_color = {
            "INFO": "#333333",
            "SUCCESS": "#388E3C",
            "WARNING": "#FFA000",
            "ERROR": "#D32F2F"
        }.get(level, "#333333")
        
        log_text.config(state=tk.NORMAL)
        log_text.insert(tk.END, f"[{timestamp}] ", "timestamp")
        log_text.insert(tk.END, f"[{level}] ", level.lower())
        log_text.insert(tk.END, f"{message}\n", "message")
        log_text.tag_config("timestamp", foreground="#666666")
        log_text.tag_config(level.lower(), foreground=level_color, font=("Arial", 9, "bold"))
        log_text.tag_config("message", foreground="#333333")
        log_text.see(tk.END)
        log_text.config(state=tk.DISABLED)
    
    return log_message

# --- GUI界面相关的函数 ---

def create_custom_button(parent, text, command, width=20, height=2):
    """创建自定义样式的按钮"""
    button = tk.Button(
        parent, 
        text=text, 
        command=command,
        width=width,
        height=height,
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        relief=tk.RAISED,
        font=("Microsoft YaHei UI", 10, "bold"),
        cursor="hand2"
    )
    return button

def create_custom_entry(parent, width=50, readonly=False):
    """创建自定义样式的输入框"""
    entry = tk.Entry(
        parent,
        width=width,
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        relief=tk.SUNKEN,
        bd=1
    )
    if readonly:
        entry.config(state='readonly')
    return entry

def create_custom_label(parent, text, width=20, anchor='w', bold=False):
    """创建自定义样式的标签"""
    font_style = ("Microsoft YaHei UI", 10, "bold" if bold else "normal")
    label = tk.Label(
        parent,
        text=text,
        width=width,
        anchor=anchor,
        font=font_style,
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR
    )
    return label

def create_custom_frame(parent, pady=5):
    """创建自定义样式的框架"""
    frame = tk.Frame(
        parent,
        pady=pady,
        bg=ACCENT_COLOR
    )
    return frame

def browse_file(entry_widget, log_func=None):
    """打开文件选择对话框，并将选定的文件路径设置到指定的Entry控件中"""
    filename = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
    )
    if filename:
        # 确保Entry控件处于可编辑状态
        entry_widget.config(state='normal')
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, filename)
        entry_widget.config(state='readonly')
        if log_func:
            log_func(f"已选择文件: {filename}", level="INFO")
    return filename

def browse_directory(entry_widget, log_func=None):
    """打开目录选择对话框，并将选定的目录路径设置到指定的Entry控件中"""
    directory = filedialog.askdirectory(title="选择输出文件夹")
    if directory:
        # 确保Entry控件处于可编辑状态
        entry_widget.config(state='normal')
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, directory)
        entry_widget.config(state='readonly')
        if log_func:
            log_func(f"已选择目录: {directory}", level="INFO")
    return directory

def open_output_folder(output_dir_path):
    """打开输出文件夹"""
    if os.path.isdir(output_dir_path):
        try:
            if os.name == 'nt':  # Windows
                os.startfile(output_dir_path)
            elif os.name == 'posix':  # macOS, Linux
                if os.uname().sysname == 'Darwin':  # macOS
                    os.system(f'open "{output_dir_path}"')
                else:  # Linux
                    os.system(f'xdg-open "{output_dir_path}"')
        except Exception as e:
            messagebox.showerror("错误", f"无法打开输出文件夹: {e}")
    else:
        messagebox.showerror("路径错误", "输出目录不存在。")

def show_about_dialog():
    """显示关于对话框"""
    # 假设 'root' 是全局可访问的主窗口实例
    # global root # 如果 root 不是全局变量而是通过其他方式传递，请相应调整

    about_window = tk.Toplevel(root) # root 应该是您Tkinter应用的主窗口实例
    about_window.title("关于")
    about_window.geometry("500x450")  # 增加了高度以容纳开发者信息
    about_window.resizable(False, False)
    about_window.transient(root)
    about_window.grab_set()

    # 设置背景色 (确保 ACCENT_COLOR 已定义)
    about_window.configure(bg=ACCENT_COLOR)

    # 标题 (确保 ACCENT_COLOR 和 PRIMARY_COLOR 已定义)
    title_label = tk.Label(
        about_window,
        text="全省基础数据表格转换与增幅计算工具",
        font=("Microsoft YaHei UI", 14, "bold"),
        bg=ACCENT_COLOR,
        fg=PRIMARY_COLOR
    )
    title_label.pack(pady=(20, 10))

    # 版本信息 (确保 ACCENT_COLOR 和 TEXT_COLOR 已定义)
    version_label = tk.Label(
        about_window,
        text="版本: 1.0.0",
        font=("Microsoft YaHei UI", 10),
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR
    )
    version_label.pack(pady=5)

    # 分隔线
    separator1 = ttk.Separator(about_window, orient='horizontal')
    separator1.pack(fill='x', padx=20, pady=10)

    # 软件说明 (确保 ACCENT_COLOR 和 TEXT_COLOR 已定义)
    description_text = """本工具用于从全省基础数据表格中提取忻州市的数据，
计算同比增幅，并生成三张Excel表格：
1. 今年数据表
2. 去年数据表
3. 包含同比增幅的对比表

适用于政务部门数据分析与报表生成。"""

    description_label = tk.Label(
        about_window,
        text=description_text,
        font=("Microsoft YaHei UI", 10),
        justify=tk.LEFT,
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR
    )
    description_label.pack(pady=10, padx=20, anchor='w')

    # --- BEGIN: Developer Information ---
    separator2 = ttk.Separator(about_window, orient='horizontal')
    separator2.pack(fill='x', padx=20, pady=(5, 10)) # pady调整

    developer_heading_label = tk.Label(
        about_window,
        text="开发信息:",
        font=("Microsoft YaHei UI", 10, "bold"),
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR, # 或 PRIMARY_COLOR
        anchor='w'
    )
    developer_heading_label.pack(padx=20, anchor='w', pady=(5,2))

    # ***** 将下面的 "[...]" 内容替换为您的真实信息 *****
    developer_info_text = """开发者: 赵越
联系邮箱: 747636400@qq.com
技术支持/单位: 忻州市烟草专卖局"""

    developer_info_label = tk.Label(
        about_window,
        text=developer_info_text,
        font=("Microsoft YaHei UI", 9),
        justify=tk.LEFT,
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR,
        anchor='w'
    )
    developer_info_label.pack(pady=(0,10), padx=20, anchor='w')
    # --- END: Developer Information ---

    # 版权信息 (确保 ACCENT_COLOR 和 TEXT_COLOR 已定义)
    # 注意：此版权信息可能需要根据开发者信息进行调整
    copyright_label = tk.Label(
        about_window,
        text="© 2025 忻州市烟草公司. 保留所有权利。",
        font=("Microsoft YaHei UI", 9),
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR
    )
    copyright_label.pack(side=tk.BOTTOM, pady=(5,10)) # pady调整以适应按钮

    # 关闭按钮 (确保 BUTTON_COLOR, BUTTON_TEXT_COLOR, HIGHLIGHT_COLOR 已定义)
    close_button = tk.Button(
        about_window,
        text="关闭",
        command=about_window.destroy,
        width=10,
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2"
    )
    close_button.pack(side=tk.BOTTOM, pady=(5,15)) # pady调整

def show_help_dialog():
    """显示帮助对话框"""
    help_window = tk.Toplevel(root)
    help_window.title("使用帮助")
    help_window.geometry("600x500")
    help_window.resizable(False, False)
    help_window.transient(root)
    help_window.grab_set()
    
    # 设置背景色
    help_window.configure(bg=ACCENT_COLOR)
    
    # 标题
    title_label = tk.Label(
        help_window, 
        text="使用帮助", 
        font=("Microsoft YaHei UI", 14, "bold"),
        bg=ACCENT_COLOR,
        fg=PRIMARY_COLOR
    )
    title_label.pack(pady=(20, 10))
    
    # 创建滚动文本区域
    help_frame = tk.Frame(help_window, bg=ACCENT_COLOR)
    help_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    scrollbar = tk.Scrollbar(help_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    help_text = tk.Text(
        help_frame, 
        wrap=tk.WORD, 
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        bd=1,
        relief=tk.SUNKEN,
        yscrollcommand=scrollbar.set
    )
    help_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=help_text.yview)
    
    # 帮助内容
    help_content = """使用指南

1. 基本操作流程
    • 选择去年的Excel文件：点击"浏览..."按钮，选择包含去年数据的Excel文件。
    • 选择今年的Excel文件：点击"浏览..."按钮，选择包含今年数据的Excel文件。
    • 选择输出目录：点击"浏览..."按钮，选择生成文件的保存位置。
    • 设置数据表前跳过行数：如果Excel文件顶部有标题或其他非数据内容，请设置需要跳过的行数。
    • 开始处理：点击"开始处理"按钮，系统将自动处理数据并生成三个Excel文件。

2. 输入文件要求
    • 文件格式：必须是Excel格式（.xlsx或.xls）。
    • 数据结构：C列必须包含市级名称，D列必须包含县级名称。
    • 指标数据：从第5列开始，每隔2列提取一个指标。

3. 输出文件说明
    • 去年数据表：包含提取的去年数据，保存为"忻州_指标数据_去年.xlsx"。
    • 今年数据表：包含提取的今年数据，保存为"忻州_指标数据_今年.xlsx"。
    • 对比表格：包含同比增幅数据，保存为"县级局分层分类基础数据.xlsx"。

4. 常见问题
    • 如果提示"未找到有效指标"，请检查Excel文件结构是否符合要求，或调整跳过行数。
    • 如果提示"未找到忻州市的数据"，请确认Excel文件中包含忻州市的数据。
    • 必须同时选择去年和今年的数据文件才能计算增幅。

5. 技术支持
    • 如遇到问题，请联系忻州市政务数据中心技术支持部门。
    • 电话：0350-XXXXXXX
    • 邮箱：support@xinzhou.gov.cn
"""
    
    help_text.insert(tk.END, help_content)
    help_text.config(state=tk.DISABLED)
    
    # 关闭按钮
    close_button = tk.Button(
        help_window, 
        text="关闭", 
        command=help_window.destroy,
        width=10,
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2"
    )
    close_button.pack(side=tk.BOTTOM, pady=15)

def run_processing():
    """获取用户在GUI上选择的各项路径和参数，然后调用核心处理函数"""
    # 从GUI获取文件路径和参数
    last_year_path = entry_last_year.get()
    current_year_path = entry_current_year.get()
    output_dir_path = entry_output_dir.get()
    
    try:
        # 获取用户输入的"跳过行数"
        rows_to_skip_val = int(entry_skip_rows.get())
        if rows_to_skip_val < 0:
            messagebox.showerror("输入错误", "跳过行数不能为负。")
            return
    except ValueError:
        messagebox.showerror("输入错误", "跳过行数必须是一个有效的数字。")
        return
    
    # 基本校验
    if not output_dir_path:
        messagebox.showerror("输入错误", "请选择输出目录。")
        status_var.set("错误: 请选择输出目录。")
        return
    if not os.path.isdir(output_dir_path):
        messagebox.showerror("路径错误", "选择的输出目录无效或不存在。")
        status_var.set("错误: 输出目录无效。")
        return
    if not last_year_path and not current_year_path:
        messagebox.showerror("输入错误", "请至少选择一个输入文件（去年或今年的数据）。")
        status_var.set("错误: 请至少选择一个输入文件。")
        return
    if not last_year_path or not current_year_path:
        messagebox.showerror("输入错误", "需要同时选择去年和今年的数据文件才能计算增幅。")
        status_var.set("错误: 请同时选择去年和今年的数据文件。")
        return
    
    # 创建进度条窗口
    progress_window = tk.Toplevel(root)
    progress_window.title("处理进度")
    progress_window.geometry("500x200")
    progress_window.resizable(False, False)
    progress_window.transient(root)
    progress_window.grab_set()
    progress_window.configure(bg=ACCENT_COLOR)
    
    # 进度标题
    progress_title = tk.Label(
        progress_window, 
        text="正在处理数据，请稍候...", 
        font=("Microsoft YaHei UI", 12, "bold"),
        bg=ACCENT_COLOR,
        fg=PRIMARY_COLOR
    )
    progress_title.pack(pady=(20, 10))
    
    # 进度详情
    progress_label = tk.Label(
        progress_window, 
        text="正在初始化...", 
        font=("Microsoft YaHei UI", 10),
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR
    )
    progress_label.pack(pady=(5, 10))
    
    # 进度条
    progress_bar = ttk.Progressbar(
        progress_window, 
        orient="horizontal", 
        length=450, 
        mode="determinate",
        style="TProgressbar"
    )
    progress_bar.pack(pady=10)
    
    # 自定义进度条样式
    style = ttk.Style()
    style.configure("TProgressbar", thickness=20, troughcolor='#E0E0E0', background=PRIMARY_COLOR)
    
    # 更新进度函数
    def update_progress(value, message):
        progress_bar["value"] = value
        progress_label.config(text=message)
        progress_window.update_idletasks()
        # 记录日志
        log_func(message)
    
    # 处理去年的文件
    update_progress(10, f"正在处理去年的文件: {os.path.basename(last_year_path)}...")
    last_year_output_file = os.path.join(output_dir_path, "忻州_指标数据_去年.xlsx")
    df_last_year, last_year_indicators = process_excel_file(last_year_path, last_year_output_file, "忻州", "去年", rows_to_skip_val, status_var, root, log_func)
    
    if df_last_year is None:
        progress_window.destroy()
        return
    
    # 处理今年的文件
    update_progress(30, f"正在处理今年的文件: {os.path.basename(current_year_path)}...")
    current_year_output_file = os.path.join(output_dir_path, "忻州_指标数据_今年.xlsx")
    df_this_year, this_year_indicators = process_excel_file(current_year_path, current_year_output_file, "忻州", "今年", rows_to_skip_val, status_var, root, log_func)
    
    if df_this_year is None:
        progress_window.destroy()
        return
    
    # 提取指标和县级市名称 (此时df_last_year和df_this_year的列名已是标准格式)
    update_progress(50, "正在提取指标和县级市名称...")
    indicators, counties = extract_indicators_and_counties(df_last_year, df_this_year)
    
    # 弹出负指标选择对话框
    update_progress(60, "请选择负指标...")
    global negative_indicators # 确保这里使用全局变量
    # Pass the root window to select_negative_indicators to make it modal correctly
    select_negative_indicators(indicators, root) # 调用函数，等待用户选择并点击确认
    log_func(f"用户已选择负指标：{', '.join(negative_indicators) if negative_indicators else '无'}", "INFO")

    # 对选定的负指标列的数值乘以-1
    update_progress(70, "正在处理负指标...")
    for indicator in negative_indicators:
        if indicator in df_this_year.columns:
            # 尝试将列转换为数值类型，无法转换的变为NaN
            df_this_year[indicator] = pd.to_numeric(df_this_year[indicator], errors='coerce')
            # 只有非NaN值才进行乘-1操作
            df_this_year[indicator] = df_this_year[indicator].apply(lambda x: x * -1 if pd.notna(x) else x)

        if indicator in df_last_year.columns:
            df_last_year[indicator] = pd.to_numeric(df_last_year[indicator], errors='coerce')
            df_last_year[indicator] = df_last_year[indicator].apply(lambda x: x * -1 if pd.notna(x) else x)
    log_func("负指标处理完成。", "INFO")

    # 计算增长率
    update_progress(80, "正在计算同比增幅...")
    result_data = calculate_growth_rates(df_last_year, df_this_year, indicators, counties)
    
    # 生成输出表格
    update_progress(90, "正在生成对比表格...")
    comparison_output_file = os.path.join(output_dir_path, "县级局分层分类基础数据.xlsx")
    df_output = generate_output_table(result_data, counties, comparison_output_file)
    
    # 添加样式
    update_progress(95, "正在为表格添加样式...")
    format_excel_with_style(comparison_output_file)
    
    # 完成处理
    update_progress(100, "处理完成！")
    
    # 关闭进度窗口
    progress_window.destroy()
    
    # 显示成功消息和结果窗口
    result_window = tk.Toplevel(root)
    result_window.title("处理成功")
    result_window.geometry("500x350")
    result_window.resizable(False, False)
    result_window.transient(root)
    result_window.grab_set()
    result_window.configure(bg=ACCENT_COLOR)
    
    # 成功图标（使用Unicode字符）
    success_icon = tk.Label(
        result_window, 
        text="✓", 
        font=("Arial", 48),
        bg=ACCENT_COLOR,
        fg=SUCCESS_COLOR
    )
    success_icon.pack(pady=(20, 10))
    
    # 成功标题
    success_title = tk.Label(
        result_window, 
        text="处理成功", 
        font=("Microsoft YaHei UI", 16, "bold"),
        bg=ACCENT_COLOR,
        fg=SUCCESS_COLOR
    )
    success_title.pack(pady=(0, 15))
    
    # 结果信息
    result_frame = tk.Frame(result_window, bg=ACCENT_COLOR)
    result_frame.pack(fill=tk.X, padx=30, pady=5)
    
    result_text = tk.Text(
        result_frame, 
        height=6, 
        width=50, 
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        bd=1,
        relief=tk.SUNKEN
    )
    result_text.pack(fill=tk.X)
    
    result_text.insert(tk.END, f"已成功生成三个文件:\n\n")
    result_text.insert(tk.END, f"1. 去年数据: {os.path.basename(last_year_output_file)}\n")
    result_text.insert(tk.END, f"2. 今年数据: {os.path.basename(current_year_output_file)}\n")
    result_text.insert(tk.END, f"3. 对比表格: {os.path.basename(comparison_output_file)}\n\n")
    result_text.insert(tk.END, f"所有文件已保存到: {output_dir_path}")
    result_text.config(state=tk.DISABLED)
    
    # 按钮区域
    button_frame = tk.Frame(result_window, bg=ACCENT_COLOR)
    button_frame.pack(fill=tk.X, padx=30, pady=20)
    
    # 打开文件夹按钮
    open_folder_button = tk.Button(
        button_frame, 
        text="打开输出文件夹", 
        command=lambda: open_output_folder(output_dir_path),
        width=15,
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2"
    )
    open_folder_button.pack(side=tk.LEFT, padx=10)
    
    # 关闭按钮
    close_button = tk.Button(
        button_frame, 
        text="关闭", 
        command=result_window.destroy,
        width=15,
        bg="#666666",
        fg=BUTTON_TEXT_COLOR,
        activebackground="#888888",
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2"
    )
    close_button.pack(side=tk.RIGHT, padx=10)
    
    # 更新状态栏
    status_var.set("所有文件处理完毕。")
    log_func("所有文件处理完毕。", level="SUCCESS")

# --- 创建主应用程序窗口 ---
if __name__ == "__main__":
    root = tk.Tk()
    root.title("忻州市烟草专卖局(公司)县级局分层分类对标数据处理系统")
    root.geometry("800x700")
    root.minsize(800, 700)
    
    # 设置应用图标（使用Unicode字符作为临时图标）
    root.iconbitmap(default=None)  # 清除默认图标
    
    # 设置窗口背景色
    root.configure(bg=ACCENT_COLOR)
    
    # 创建主框架
    main_frame = tk.Frame(root, bg=ACCENT_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    # --- 顶部标题区域 ---
    header_frame = tk.Frame(main_frame, bg=PRIMARY_COLOR, height=80)
    header_frame.pack(fill=tk.X, pady=(0, 20))
    
    # 应用标题
    title_label = tk.Label(
        header_frame, 
        text="忻州市烟草专卖局(公司)县级局分层分类对标数据处理系统", 
        font=("Microsoft YaHei UI", 12, "bold"),
        bg=PRIMARY_COLOR,
        fg="white",
        pady=20
    )
    title_label.pack(side=tk.LEFT, padx=30)
    
    # 当前日期
    current_date = datetime.datetime.now().strftime("%Y年%m月%d日")
    date_label = tk.Label(
        header_frame, 
        text=current_date, 
        font=("Microsoft YaHei UI", 10),
        bg=PRIMARY_COLOR,
        fg="white"
    )
    date_label.pack(side=tk.RIGHT, padx=30, pady=20)
    
    # --- 菜单栏 ---
    menu_bar = tk.Menu(root)
    root.config(menu=menu_bar)
    
    # 文件菜单
    file_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="文件", menu=file_menu)
    file_menu.add_command(label="开始处理", command=run_processing)
    file_menu.add_separator()
    file_menu.add_command(label="退出", command=root.quit)
    
    # 帮助菜单
    help_menu = tk.Menu(menu_bar, tearoff=0)
    menu_bar.add_cascade(label="帮助", menu=help_menu)
    help_menu.add_command(label="使用帮助", command=show_help_dialog)
    help_menu.add_command(label="关于", command=show_about_dialog)
    
    # --- 主内容区域 ---
    content_frame = tk.Frame(main_frame, bg=ACCENT_COLOR)
    content_frame.pack(fill=tk.BOTH, expand=True, pady=10)
    
    # 左侧面板
    left_panel = tk.Frame(content_frame, bg=ACCENT_COLOR, width=200)
    left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 20))
    
    # 左侧面板标题
    panel_title = tk.Label(
        left_panel, 
        text="操作面板", 
        font=("Microsoft YaHei UI", 12, "bold"),
        bg=PRIMARY_COLOR,
        fg="white",
        width=20,
        pady=10
    )
    panel_title.pack(fill=tk.X)
    
    # 左侧面板按钮
    process_button = create_custom_button(left_panel, "开始处理", run_processing)
    process_button.pack(fill=tk.X, pady=(20, 10))
    
    help_button = create_custom_button(left_panel, "使用帮助", show_help_dialog)
    help_button.pack(fill=tk.X, pady=10)
    
    about_button = create_custom_button(left_panel, "关于系统", show_about_dialog)
    about_button.pack(fill=tk.X, pady=10)
    
    exit_button = create_custom_button(left_panel, "退出系统", root.quit)
    exit_button.pack(fill=tk.X, pady=10)
    
    # 右侧内容区
    right_content = tk.Frame(content_frame, bg="white", bd=1, relief=tk.SOLID)
    right_content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
    
    # 内容区标题
    content_title = tk.Label(
        right_content, 
        text="数据处理配置", 
        font=("Microsoft YaHei UI", 12, "bold"),
        bg=SECONDARY_COLOR,
        fg="white",
        pady=10
    )
    content_title.pack(fill=tk.X)
    
    # 配置区域
    config_frame = tk.Frame(right_content, bg="white", padx=20, pady=20)
    config_frame.pack(fill=tk.BOTH, expand=True)
    
    # --- "数据表前跳过行数" 输入区域 ---
    frame_skip_rows = create_custom_frame(config_frame)
    frame_skip_rows.pack(fill=tk.X, pady=(0, 15))
    
    lbl_skip_rows = create_custom_label(frame_skip_rows, "数据表前跳过行数:", width=18, bold=True)
    lbl_skip_rows.pack(side=tk.LEFT)
    
    entry_skip_rows = create_custom_entry(frame_skip_rows, width=10)
    entry_skip_rows.insert(0, str(ROWS_TO_SKIP_DEFAULT))
    entry_skip_rows.pack(side=tk.LEFT, padx=5)
    
    hint_label = tk.Label(
        frame_skip_rows, 
        text="(C列为市的数据表内容前的总行数)", 
        font=("Microsoft YaHei UI", 9, "italic"),
        bg=ACCENT_COLOR,
        fg="#666666"
    )
    hint_label.pack(side=tk.LEFT, padx=5)
    
    # --- 日志区域 ---
    log_frame = tk.LabelFrame(
        main_frame, 
        text="操作日志", 
        font=("Microsoft YaHei UI", 10, "bold"),
        bg=ACCENT_COLOR,
        fg=TEXT_COLOR,
        padx=10,
        pady=10
    )
    log_frame.pack(fill=tk.X, padx=20, pady=(10, 20))
    
    # 创建日志文本区域
    log_text = tk.Text(
        log_frame, 
        height=8, 
        wrap=tk.WORD, 
        font=("Consolas", 9),
        bg="white",
        fg=TEXT_COLOR,
        bd=1,
        relief=tk.SUNKEN
    )
    log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # 添加滚动条
    log_scrollbar = tk.Scrollbar(log_frame)
    log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # 连接滚动条和文本区域
    log_text.config(yscrollcommand=log_scrollbar.set)
    log_scrollbar.config(command=log_text.yview)
    
    # 初始化日志区域
    log_text.config(state=tk.DISABLED)
    
    # 创建日志函数
    log_func = create_log_function(log_text)
    
    # 初始日志
    log_func("系统已启动，等待用户操作...", level="INFO")
    
    # --- 文件选择区域 ---
    # 去年文件
    frame_last_year = create_custom_frame(config_frame)
    frame_last_year.pack(fill=tk.X, pady=10)
    
    lbl_last_year = create_custom_label(frame_last_year, "去年Excel文件:", width=18, bold=True)
    lbl_last_year.pack(side=tk.LEFT)
    
    # 创建一个容器框架来包含输入框和按钮
    last_year_container = tk.Frame(frame_last_year, bg=ACCENT_COLOR)
    last_year_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    # 输入框
    entry_last_year = tk.Entry(
        last_year_container,
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        relief=tk.SUNKEN,
        bd=1
    )
    entry_last_year.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    
    # 浏览按钮 - 直接使用tk.Button而不是自定义函数
    btn_last_year = tk.Button(
        last_year_container,
        text="浏览...",
        command=lambda: browse_file(entry_last_year, log_func),
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2",
        relief=tk.RAISED,
        width=8
    )
    btn_last_year.pack(side=tk.RIGHT)
    
    # 今年文件
    frame_current_year = create_custom_frame(config_frame)
    frame_current_year.pack(fill=tk.X, pady=10)
    
    lbl_current_year = create_custom_label(frame_current_year, "今年Excel文件:", width=18, bold=True)
    lbl_current_year.pack(side=tk.LEFT)
    
    # 创建一个容器框架来包含输入框和按钮
    current_year_container = tk.Frame(frame_current_year, bg=ACCENT_COLOR)
    current_year_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    # 输入框
    entry_current_year = tk.Entry(
        current_year_container,
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        relief=tk.SUNKEN,
        bd=1
    )
    entry_current_year.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    
    # 浏览按钮
    btn_current_year = tk.Button(
        current_year_container,
        text="浏览...",
        command=lambda: browse_file(entry_current_year, log_func),
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2",
        relief=tk.RAISED,
        width=8
    )
    btn_current_year.pack(side=tk.RIGHT)
    
    # 输出目录
    frame_output_dir = create_custom_frame(config_frame)
    frame_output_dir.pack(fill=tk.X, pady=10)
    
    lbl_output_dir = create_custom_label(frame_output_dir, "输出目录:", width=18, bold=True)
    lbl_output_dir.pack(side=tk.LEFT)
    
    # 创建一个容器框架来包含输入框和按钮
    output_dir_container = tk.Frame(frame_output_dir, bg=ACCENT_COLOR)
    output_dir_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    # 输入框
    entry_output_dir = tk.Entry(
        output_dir_container,
        font=("Microsoft YaHei UI", 10),
        bg="white",
        fg=TEXT_COLOR,
        relief=tk.SUNKEN,
        bd=1
    )
    entry_output_dir.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    
    # 浏览按钮
    btn_output_dir = tk.Button(
        output_dir_container,
        text="浏览...",
        command=lambda: browse_directory(entry_output_dir, log_func),
        bg=BUTTON_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground=HIGHLIGHT_COLOR,
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 10),
        cursor="hand2",
        relief=tk.RAISED,
        width=8
    )
    btn_output_dir.pack(side=tk.RIGHT)
    
    # --- 说明文本 ---
    info_frame = tk.Frame(config_frame, bg="white", pady=15)
    info_frame.pack(fill=tk.X)
    
    info_text = tk.Label(
        info_frame, 
        text="说明: 本工具将生成三个文件 - 去年数据表、今年数据表和包含同比增幅的对比表。\n需要同时选择去年和今年的数据文件才能计算增幅。", 
        font=("Microsoft YaHei UI", 9, "italic"),
        justify=tk.LEFT,
        bg="white",
        fg="#666666"
    )
    info_text.pack(anchor='w')
    
    # 添加一个明显的"开始处理"按钮在配置区域底部
    process_button_main = tk.Button(
        config_frame,
        text="开始处理",
        command=run_processing,
        bg=SUCCESS_COLOR,
        fg=BUTTON_TEXT_COLOR,
        activebackground="#2E7D32",
        activeforeground=BUTTON_TEXT_COLOR,
        font=("Microsoft YaHei UI", 12, "bold"),
        cursor="hand2",
        relief=tk.RAISED,
        width=20,
        height=2
    )
    process_button_main.pack(pady=20)
    
    # --- 状态栏 ---
    status_var = tk.StringVar()
    status_var.set("就绪：请选择文件并点击开始处理。")
    status_frame = tk.Frame(root, bg=PRIMARY_COLOR, height=25)
    status_frame.pack(side=tk.BOTTOM, fill=tk.X)
    
    status_label = tk.Label(
        status_frame, 
        textvariable=status_var, 
        font=("Microsoft YaHei UI", 9),
        bg=PRIMARY_COLOR,
        fg="white",
        bd=1,
        anchor='w',
        padx=10,
        pady=5
    )
    status_label.pack(side=tk.LEFT, fill=tk.X)
    
    # 版权信息
    copyright_label = tk.Label(
        status_frame, 
        text="© 2025 忻州市政务数据中心", 
        font=("Microsoft YaHei UI", 8),
        bg=PRIMARY_COLOR,
        fg="white",
        padx=10,
        pady=5
    )
    copyright_label.pack(side=tk.RIGHT)
    
    # 启动主循环
    root.mainloop()
    print("1")
    